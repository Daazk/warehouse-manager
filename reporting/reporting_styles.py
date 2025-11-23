from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADER_FONT = Font(size=12, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill("solid", fgColor="FF1F497D")
THIN_BORDER = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin"),
)

def apply_header_style(ws, columns, widths=None):
    """Применяет стиль к заголовкам указанного листа."""
    
    for col in columns:
        cell = ws[f"{col}1"]
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL

        if widths and col in widths:
            ws.column_dimensions[col].width = widths[col]


def apply_table_style(ws):
    """Выравнивание и рамки для всей таблицы."""

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CENTER
            cell.border = THIN_BORDER