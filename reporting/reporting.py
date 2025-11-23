from openpyxl import Workbook
from utils import REPORT
from .reporting_styles import apply_header_style, apply_table_style

def create_workbook():
    """Функия создания excel файла для отчёта"""

    wb = Workbook()
    wb.remove(wb.active)
    return wb

def add_sklad_sheet(wb, sklad):
    """Отчёт по складу"""

    ws = wb.create_sheet("Склад")

    # Заголовки
    headers = ["ID", "Название", "Категория", "Цена", "Количество"]
    ws.append(headers)

    # Стиль заголовков
    apply_header_style(
        ws,
        columns=["A", "B", "C", "D", "E"],
        widths={"A": 8, "B": 30, "C": 30, "D": 12, "E": 14}
    )

    # Данные склада
    for pid, item in sklad.items():
        ws.append([pid, item["name"], item["category"], item["price"], item["quantity"]])

    # Общий стиль таблицы
    apply_table_style(ws)

def add_clients_sheet(wb, clients):
    """Отчёт по клиентам"""

    ws = wb.create_sheet("Клиенты")

    headers = ["ID", "Имя", "Телефон", "Email"]
    ws.append(headers)

    apply_header_style(
        ws,
        columns=["A", "B", "C", "D"],
        widths={"A": 8, "B": 25, "C": 20, "D": 30}
    )

    for cid, data in clients.items():
        ws.append([cid, data["name"], data["phone"], data["email"]])

    apply_table_style(ws)

def add_orders_sheet(wb, orders):
    """Отчёт по заказам"""

    ws = wb.create_sheet("Заказы")

    headers = ["ID", "Клиент", "Дата", "Сумма", "Статус"]
    ws.append(headers)

    apply_header_style(
        ws,
        columns=["A", "B", "C", "D", "E"],
        widths={"A": 8, "B": 11, "C": 20, "D": 30, "E": 30}
    )

    for cid, data in orders.items():
        ws.append([cid, data["client_id"], data["date"], data["total"], data["status"]])

    apply_table_style(ws)


def generate_report(sklad=None, clients=None, orders=None, filename=REPORT):
    """Генерация отчёта"""

    wb = create_workbook()

    if sklad:
        add_sklad_sheet(wb, sklad)

    if clients:
        add_clients_sheet(wb, clients)

    if orders:
        add_orders_sheet(wb, orders)

    wb.save(filename)
    return filename
